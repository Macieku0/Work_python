import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import pandas
import os
import win32api
from win32com import client

excel = client.DispatchEx("Excel.Application")
excel.Visible = 0

#Main files directory
MainDir = "C:/Users/macie/Pulpit/III_CAT_SPRAWDZENIE_SRUB/20210519_LISTA_WPALEN_python/"
#List name
ListName = "LISTA_WPALEK.xlsx"
#Name of list worksheet
WorkSheetName = "Tie-In list"
#Letter of column with names for naming files
Col_lett = "K"
#Range of column with names of files
ColStart = 12
ColEnd = 15
#Template name
TemplateName = "Karta technologiczna_wpalki.xlsx"
#Name of template worksheet 
TemplateWorkSheetName = "KARTA"
#Files save / update directory
FilesDir = "C:/Users/macie/Pulpit/III_CAT_SPRAWDZENIE_SRUB/20210519_LISTA_WPALEN_python/WYDRUKI/"
#List of files to create / update
FilesList = []
#Change color marker
MyColor = "E7CE63"
#Map of cells connection between list and template
#TODO Map rest of the cells
Map = {'A7':'C','B7':'G','C7':'D'}

#Creating a list of all files name
source = openpyxl.load_workbook(MainDir + TemplateName)
wb = openpyxl.load_workbook(MainDir + ListName)
worksheet = wb[WorkSheetName]
i = 0
for z in range(ColStart,ColEnd):
    file = worksheet[Col_lett + str(z)].value
    #Check if file already exist
    if os.path.isfile(FilesDir + file + ".xlsx"):
        print(file + ".xlsx" + " already exist")
        NewFile = openpyxl.load_workbook(FilesDir + file + ".xlsx")
        NewWorksheet = NewFile[TemplateWorkSheetName]
        for x,y in Map.items():
            if NewWorksheet[x].value != worksheet[y + str(z)].value:
                comment = Comment('Previous value = ' + str(NewWorksheet[x].value),'automatic inspect')
                NewWorksheet[x].value = worksheet[y + str(z)].value
                NewWorksheet[x].fill = PatternFill(fgColor=MyColor, fill_type="solid")
                NewWorksheet[x].comment = comment
                NewFile.save(FilesDir + file + ".xlsx")
                path = str(FilesDir + file + ".pdf")
                WbPrint = excel.Workbooks.Open(FilesDir + file + ".xlsx")
                WsPrint = WbPrint.Worksheets[TemplateWorkSheetName]
                WbPrint.SaveAs(path,FileFormat=57)
                WbPrint.Close()
                excel.Quit()
                NewFile.close()
    else:
        #If no so create one
        source.save(FilesDir + file + ".xlsx")
        source.close()
        #Fullfil created files with data from list
        NewFile = openpyxl.load_workbook(FilesDir + file + ".xlsx")
        NewWorksheet = NewFile[TemplateWorkSheetName]
        for x,y in Map.items():
            NewWorksheet[x].value = worksheet[y + str(z)].value
            NewFile.save(FilesDir + file + ".xlsx")
            path = str(FilesDir + file + ".pdf")
            WbPrint = excel.Workbooks.Open(FilesDir + file + ".xlsx")
            WsPrint = WbPrint.Worksheets[TemplateWorkSheetName]
            WbPrint.SaveAs(path,FileFormat=57)
            WbPrint.Close()
            excel.Quit()
            NewFile.close()
    i += 1
    #Close list
wb.close()

